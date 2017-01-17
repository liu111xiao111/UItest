import os
import time
import datetime
import shutil
import zipfile 
import openpyxl
from tools.utility.constants import *


ANR_ERROR = 0
JNT_ERROR = 0
JNT_CRASH = 0
APP_DIED = 0
SYSTEM_ERROR = 0

class DataHandler(object):
    def __init__(self):
        self.rsPath = ''
        self.testCase = ''

    def handle(self, rsPath):
        self.rsPath = rsPath
        self.reportPath = os.path.join(self.rsPath, 'attachment')
        self.logPath = os.path.join(self.reportPath, 'log')

        self._summaryLogData()
        self._addZipFile()

    def getLogData(self,testCase):
        self.testCase = testCase
        return self._parserLogData()

    #把Log文件信息打包 
    def _addZipFile(self):
        zipf = zipfile.ZipFile('log.zip', 'w')
        pre_len = len(os.path.dirname(self.logPath))
        for parent, dirnames, filenames in os.walk(self.logPath):
            for filename in filenames:
                pathfile = os.path.join(parent, filename)
                arcname = pathfile[pre_len:].strip(os.path.sep)
                zipf.write(pathfile, arcname)
        zipf.close()
        zipFileName = os.path.join(self.reportPath, 'log.zip')
        shutil.move("log.zip", zipFileName)

    def _summaryLogData(self):
        if not os.path.exists(self.logPath):
            os.makedirs(self.logPath)

        cmdCP = "cp -R %s/[0-9] %s" % (self.rsPath, self.logPath)
        os.system(cmdCP)

        tmpFile = os.path.join(self.logPath, 'tmp.txt')
        cmdFind = 'find %s -type d -name "screenshot" > %s' % (self.logPath, tmpFile)
        os.system(cmdFind)
        rmPath = open(tmpFile, 'r')
        while 1:
            line = rmPath.readline()
            if not line:
                break
            else:
                cmdRM = "rm -rf %s" % (line)
                os.system(cmdRM)
        if os.path.exists(tmpFile):
            os.remove(tmpFile)

    def _parserLogData(self):
        i = 0
        caseErrorInfo = {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]}
        tmpFile = os.path.join(self.logPath, '%s.txt') % self.testCase
        cmdFind = 'find %s -name "%s*.log" > %s' % (self.logPath, self.testCase, tmpFile)
        os.system(cmdFind)
        if os.path.exists(tmpFile):
            logCasePath = open(tmpFile, 'r')
            logPaths = logCasePath.readlines()
            logCasePath.close()
            if logPaths != []:
                for logPath in logPaths:
                    caseLogFile = logPath[:-1]
                    errorFilePath = caseLogFile.find('log')
                    if os.path.exists(caseLogFile):
                        logInfo = open(caseLogFile, 'r')
                        logLines = logInfo.readlines()
                        logInfo.close()
                        for logLine in logLines:
                            i += 1
                            if logLine.find("anr") != -1:
                                caseErrorInfo['ANR'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                            elif logLine.find("crash") != -1:
                                caseErrorInfo['JRTCRASH'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                            elif logLine.find("Reading a NULL string not supported here.") != -1:
                                caseErrorInfo['JRTERROR'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                            elif logLine.find("Got null root node from accessibility - Retrying...") != -1:
                                caseErrorInfo['JRTERROR'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                            elif logLine.find("died") != -1:
                                caseErrorInfo['APPDIED'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                            elif logLine.find("system error") != -1:
                                caseErrorInfo['SYSTEMERROR'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
        if os.path.exists(tmpFile):
            os.remove(tmpFile)
        return caseErrorInfo


class Handler(object):
    def __init__(self, deviceType):
        self.rsPath = ''
        self.reportPath = ''
        self.workbook = ''
        self.dataLength = 0
        self.dataList = {u'quanchengsousuo': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]}, 
                         u'gouwuzhongxin': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'meishihui': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangsousuo': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangzhaodian': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangpaidui': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangtingche': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangmaidan': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'wodedenglu': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'wodetuichu': {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]}}
        if deviceType == 'Android':
            from configs.androidConfig import appVersion, phoneVersion, buildVersion, deviceID, deviceNet
        elif deviceType == 'IOS':
            from configs.iosConfig import appVersion, phoneVersion, buildVersion, deviceID, deviceNet
        else:
            raise

        self.deviceType = deviceType
        self.appVersion = appVersion
        self.phoneVersion = phoneVersion
        self.buildVersion = buildVersion
        self.deviceID = deviceID
        self.deviceNet = deviceNet

    def handle(self, rsPath):
        self.rsPath = rsPath
        self._mkReportDir()
        try:
            resourcesDirectory = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/resources/"
            xlsFileTemplate = os.path.join(resourcesDirectory, u'templateStability.xlsx')
            shutil.copy(xlsFileTemplate,self.reportPath)
            dataHandler = DataHandler()
            dataHandler.handle(rsPath=self.rsPath)
            for testCase in (STABILITY_CASE_FOLDER_LIST.values()):
                caseErrorInfo = dataHandler.getLogData(testCase=testCase)
                self.dataList[testCase]['ANR'] = caseErrorInfo['ANR']
                self.dataList[testCase]['JRTERROR'] = caseErrorInfo['JRTERROR']
                self.dataList[testCase]['JRTCRASH'] = caseErrorInfo['JRTCRASH']
                self.dataList[testCase]['APPDIED'] = caseErrorInfo['APPDIED']
                self.dataList[testCase]['SYSTEMERROR'] = caseErrorInfo['SYSTEMERROR']
                self.dataList[testCase]['ERRORLOG'] = caseErrorInfo['ERRORLOG']
                self.dataList[testCase]['ERRORNUM'] = caseErrorInfo['ERRORNUM']
                self.dataList[testCase]['ERRORFILE'] = caseErrorInfo['ERRORFILE']

            global ANR_ERROR
            ANR_ERROR = self.dataList['quanchengsousuo']['ANR'] + self.dataList['gouwuzhongxin']['ANR'] +\
                        self.dataList['meishihui']['ANR'] + self.dataList['guangchangsousuo']['ANR'] + \
                        self.dataList['guangchangzhaodian']['ANR'] + self.dataList['guangchangpaidui']['ANR'] +\
                        self.dataList['guangchangtingche']['ANR'] + self.dataList['guangchangmaidan']['ANR'] +\
                        self.dataList['wodedenglu']['ANR'] + self.dataList['wodetuichu']['ANR']
            global JNT_ERROR
            JNT_ERROR = self.dataList['quanchengsousuo']['JRTERROR'] + self.dataList['gouwuzhongxin']['JRTERROR'] +\
                        self.dataList['meishihui']['JRTERROR'] + self.dataList['guangchangsousuo']['JRTERROR'] +\
                        self.dataList['guangchangzhaodian']['JRTERROR'] + self.dataList['guangchangpaidui']['JRTERROR'] +\
                        self.dataList['guangchangtingche']['JRTERROR'] + self.dataList['guangchangmaidan']['JRTERROR'] +\
                        self.dataList['wodedenglu']['JRTERROR'] + self.dataList['wodetuichu']['JRTERROR']
            global JNT_CRASH
            JNT_CRASH = self.dataList['quanchengsousuo']['JRTCRASH'] + self.dataList['gouwuzhongxin']['JRTCRASH'] +\
                        self.dataList['meishihui']['JRTCRASH'] + self.dataList['guangchangsousuo']['JRTCRASH'] +\
                        self.dataList['guangchangzhaodian']['JRTCRASH'] + self.dataList['guangchangpaidui']['JRTCRASH'] +\
                        self.dataList['guangchangtingche']['JRTCRASH'] + self.dataList['guangchangmaidan']['JRTCRASH'] +\
                        self.dataList['wodedenglu']['JRTCRASH'] + self.dataList['wodetuichu']['JRTCRASH']
            global APP_DIED
            APP_DIED = self.dataList['quanchengsousuo']['APPDIED'] + self.dataList['gouwuzhongxin']['APPDIED'] +\
                        self.dataList['meishihui']['APPDIED'] + self.dataList['guangchangsousuo']['APPDIED'] +\
                        self.dataList['guangchangzhaodian']['APPDIED'] + self.dataList['guangchangpaidui']['APPDIED'] +\
                        self.dataList['guangchangtingche']['APPDIED'] + self.dataList['guangchangmaidan']['APPDIED'] +\
                        self.dataList['wodedenglu']['APPDIED'] + self.dataList['wodetuichu']['APPDIED']
            global SYSTEM_ERROR
            SYSTEM_ERROR = self.dataList['quanchengsousuo']['SYSTEMERROR'] + self.dataList['gouwuzhongxin']['SYSTEMERROR'] +\
                           self.dataList['meishihui']['SYSTEMERROR'] + self.dataList['guangchangsousuo']['SYSTEMERROR'] + \
                           self.dataList['guangchangzhaodian']['SYSTEMERROR'] + self.dataList['guangchangpaidui']['SYSTEMERROR'] +\
                           self.dataList['guangchangtingche']['SYSTEMERROR'] + self.dataList['guangchangmaidan']['SYSTEMERROR'] +\
                           self.dataList['wodedenglu']['SYSTEMERROR'] + self.dataList['wodetuichu']['SYSTEMERROR']
            xlsFile = os.path.join(self.reportPath, u'templateStability.xlsx')
        except Exception as e:
            print(e)
        finally:
            self._writeExcel(xlsFile)

    def _mkReportDir(self):
        '''
        创建报告目录, 包含excel报告,log压缩包
        '''
        self.reportPath = os.path.join(self.rsPath, 'attachment')
        if os.path.exists(self.reportPath):
            timestamp = os.path.getmtime(self.reportPath)
            date = datetime.datetime.fromtimestamp(timestamp).strftime('%Y_%m_%d_%H_%M_%S')
            newDirName = 'report_' + date

            newLogDir = os.path.join(self.rsPath, newDirName)
            try:
                os.renames(self.reportPath, newLogDir)
            except Exception as e:
                raise IOError('Modify the [%s] directory name failed with following error: \n'
                              '%s' % (self.reportPath, e))

        try:
            os.makedirs(self.reportPath)
        except Exception as e:
            raise IOError('Create the [%s] directory failed with following error: \n'
                          '%s' % (self.reportPath, e))

    def _writeExcel(self, file='file.xls'):
        try:
            wb = openpyxl.load_workbook(file)
            for sheet_name in STABILITY_REPORT_SHEET:
                ws = wb.get_sheet_by_name(sheet_name)
                if sheet_name == u'测试环境':
                    ws['C4'] = self.phoneVersion
                    ws['C5'] = self.buildVersion
                    ws['C9'] = self.appVersion
                    ws['A13'] = self.deviceNet
                elif sheet_name == u'全城搜索':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['quanchengsousuo']['ANR'] + self.dataList['quanchengsousuo']['JRTERROR'] + self.dataList['quanchengsousuo']['JRTCRASH'] + self.dataList['quanchengsousuo']['APPDIED'] + self.dataList['quanchengsousuo']['SYSTEMERROR']
                    ws['C6'] = self.dataList['quanchengsousuo']['ANR']
                    ws['C7'] = self.dataList['quanchengsousuo']['JRTERROR'] + self.dataList['quanchengsousuo']['JRTCRASH']
                    ws['C8'] = self.dataList['quanchengsousuo']['APPDIED'] + self.dataList['quanchengsousuo']['SYSTEMERROR']
                    if len(self.dataList['quanchengsousuo']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['quanchengsousuo']['ERRORLOG'])):
                            log_quanchengsousuo = 'A' + str(i + 11)
                            line_quanchengsousuo = 'E' + str(i + 11)
                            message_quanchengsousuo = 'F' + str(i + 11)
                            ws[line_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORNUM'][i]
                            ws[message_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORLOG'][i]
                            ws[log_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'购物中心':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['gouwuzhongxin']['ANR'] + self.dataList['gouwuzhongxin']['JRTERROR'] + self.dataList['gouwuzhongxin']['JRTCRASH'] + self.dataList['gouwuzhongxin']['APPDIED'] + self.dataList['gouwuzhongxin']['SYSTEMERROR']
                    ws['C6'] = self.dataList['gouwuzhongxin']['ANR']
                    ws['C7'] = self.dataList['gouwuzhongxin']['JRTERROR'] + self.dataList['gouwuzhongxin']['JRTCRASH']
                    ws['C8'] = self.dataList['gouwuzhongxin']['APPDIED'] + self.dataList['gouwuzhongxin']['SYSTEMERROR']
                    if len(self.dataList['gouwuzhongxin']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['gouwuzhongxin']['ERRORLOG'])):
                            log_gouwuzhongxin = 'A' + str(i + 11)
                            line_gouwuzhongxin = 'E' + str(i + 11)
                            message_gouwuzhongxin = 'F' + str(i + 11)
                            ws[line_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORNUM'][i]
                            ws[message_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORLOG'][i]
                            ws[log_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'美食汇':
                    ws['C4'] = OUTLOOPNUM
                    ws['C5'] = self.dataList['meishihui']['ANR'] + self.dataList['meishihui']['JRTERROR'] + self.dataList['meishihui']['JRTCRASH'] + self.dataList['meishihui']['APPDIED'] + self.dataList['meishihui']['SYSTEMERROR']
                    ws['C6'] = self.dataList['meishihui']['ANR']
                    ws['C7'] = self.dataList['meishihui']['JRTERROR'] + self.dataList['meishihui']['JRTCRASH']
                    ws['C8'] = self.dataList['meishihui']['APPDIED'] + self.dataList['meishihui']['SYSTEMERROR']
                    if len(self.dataList['meishihui']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['meishihui']['ERRORLOG'])):
                            log_meishihui = 'A' + str(i + 11)
                            line_meishihui = 'E' + str(i + 11)
                            message_meishihui = 'F' + str(i + 11)
                            ws[line_meishihui] = self.dataList['meishihui']['ERRORNUM'][i]
                            ws[message_meishihui] = self.dataList['meishihui']['ERRORLOG'][i]
                            ws[log_meishihui] = self.dataList['meishihui']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场搜索':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangsousuo']['ANR'] + self.dataList['guangchangsousuo']['JRTERROR'] + self.dataList['guangchangsousuo']['JRTCRASH'] + self.dataList['guangchangsousuo']['APPDIED'] + self.dataList['guangchangsousuo']['SYSTEMERROR']
                    ws['C6'] = self.dataList['guangchangsousuo']['ANR']
                    ws['C7'] = self.dataList['guangchangsousuo']['JRTERROR'] + self.dataList['guangchangsousuo']['JRTCRASH']
                    ws['C8'] = self.dataList['guangchangsousuo']['APPDIED'] + self.dataList['guangchangsousuo']['SYSTEMERROR']
                    if len(self.dataList['guangchangsousuo']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangsousuo']['ERRORLOG'])):
                            log_guangchangsousuo = 'A' + str(i + 11)
                            line_guangchangsousuo = 'E' + str(i + 11)
                            message_guangchangsousuo = 'F' + str(i + 11)
                            ws[line_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORNUM'][i]
                            ws[message_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORLOG'][i]
                            ws[log_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场找店':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangzhaodian']['ANR'] + self.dataList['guangchangzhaodian']['JRTERROR'] + self.dataList['guangchangzhaodian']['JRTCRASH'] + self.dataList['guangchangzhaodian']['APPDIED'] + self.dataList['guangchangzhaodian']['SYSTEMERROR']
                    ws['C6'] = self.dataList['guangchangzhaodian']['ANR']
                    ws['C7'] = self.dataList['guangchangzhaodian']['JRTERROR'] + self.dataList['guangchangzhaodian']['JRTCRASH']
                    ws['C8'] = self.dataList['guangchangzhaodian']['APPDIED'] + self.dataList['guangchangzhaodian']['SYSTEMERROR']
                    if len(self.dataList['guangchangzhaodian']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangzhaodian']['ERRORLOG'])):
                            log_guangchangzhaodian = 'A' + str(i + 11)
                            line_guangchangzhaodian = 'E' + str(i + 11)
                            message_guangchangzhaodian = 'F' + str(i + 11)
                            ws[line_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORNUM'][i]
                            ws[message_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORLOG'][i]
                            ws[log_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场排队':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangpaidui']['ANR'] + self.dataList['guangchangpaidui']['JRTERROR'] + self.dataList['guangchangpaidui']['JRTCRASH'] + self.dataList['guangchangpaidui']['APPDIED'] + self.dataList['guangchangpaidui']['SYSTEMERROR']
                    ws['C6'] = self.dataList['guangchangpaidui']['ANR']
                    ws['C7'] = self.dataList['guangchangpaidui']['JRTERROR'] + self.dataList['guangchangpaidui']['JRTCRASH']
                    ws['C8'] = self.dataList['guangchangpaidui']['APPDIED'] + self.dataList['guangchangpaidui']['SYSTEMERROR']
                    if len(self.dataList['guangchangpaidui']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangpaidui']['ERRORLOG'])):
                            log_guangchangpaidui = 'A' + str(i + 11)
                            line_guangchangpaidui = 'E' + str(i + 11)
                            message_guangchangpaidui = 'F' + str(i + 11)
                            ws[line_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORNUM'][i]
                            ws[message_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORLOG'][i]
                            ws[log_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场停车':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangtingche']['ANR'] + self.dataList['guangchangtingche']['JRTERROR'] + self.dataList['guangchangtingche']['JRTCRASH'] + self.dataList['guangchangtingche']['APPDIED'] + self.dataList['guangchangtingche']['SYSTEMERROR']
                    ws['C6'] = self.dataList['guangchangtingche']['ANR']
                    ws['C7'] = self.dataList['guangchangtingche']['JRTERROR'] + self.dataList['guangchangtingche']['JRTCRASH']
                    ws['C8'] = self.dataList['guangchangtingche']['APPDIED'] + self.dataList['guangchangtingche']['SYSTEMERROR']
                    if len(self.dataList['guangchangtingche']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangtingche']['ERRORLOG'])):
                            log_guangchangtingche = 'A' + str(i + 11)
                            line_guangchangtingche = 'E' + str(i + 11)
                            message_guangchangtingche = 'F' + str(i + 11)
                            ws[line_guangchangtingche] = self.dataList['guangchangtingche']['ERRORNUM'][i]
                            ws[message_guangchangtingche] = self.dataList['guangchangtingche']['ERRORLOG'][i]
                            ws[log_guangchangtingche] = self.dataList['guangchangtingche']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场买单':
                    ws['C4'] = OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangmaidan']['ANR'] + self.dataList['guangchangmaidan']['JRTERROR'] + self.dataList['guangchangmaidan']['JRTCRASH'] + self.dataList['guangchangmaidan']['APPDIED'] + self.dataList['guangchangmaidan']['SYSTEMERROR']
                    ws['C6'] = self.dataList['guangchangmaidan']['ANR']
                    ws['C7'] = self.dataList['guangchangmaidan']['JRTERROR'] + self.dataList['guangchangmaidan']['JRTCRASH']
                    ws['C8'] = self.dataList['guangchangmaidan']['APPDIED'] + self.dataList['guangchangmaidan']['SYSTEMERROR']
                    if len(self.dataList['guangchangmaidan']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangmaidan']['ERRORLOG'])):
                            log_guangchangmaidan = 'A' + str(i + 11)
                            line_guangchangmaidan = 'E' + str(i + 11)
                            message_guangchangmaidan = 'F' + str(i + 11)
                            ws[line_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORNUM'][i]
                            ws[message_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORLOG'][i]
                            ws[log_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'我的登录':
                    ws['C4'] = OUTLOOPNUM
                    ws['C5'] = self.dataList['wodedenglu']['ANR'] + self.dataList['wodedenglu']['JRTERROR'] + self.dataList['wodedenglu']['JRTCRASH'] + self.dataList['wodedenglu']['APPDIED'] + self.dataList['wodedenglu']['SYSTEMERROR']
                    ws['C6'] = self.dataList['wodedenglu']['ANR']
                    ws['C7'] = self.dataList['wodedenglu']['JRTERROR'] + self.dataList['wodedenglu']['JRTCRASH']
                    ws['C8'] = self.dataList['wodedenglu']['APPDIED'] + self.dataList['wodedenglu']['SYSTEMERROR']
                    if len(self.dataList['wodedenglu']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['wodedenglu']['ERRORLOG'])):
                            log_wodedenglu = 'A' + str(i + 11)
                            line_wodedenglu = 'E' + str(i + 11)
                            message_wodedenglu = 'F' + str(i + 11)
                            ws[line_wodedenglu] = self.dataList['wodedenglu']['ERRORNUM'][i]
                            ws[message_wodedenglu] = self.dataList['wodedenglu']['ERRORLOG'][i]
                            ws[log_wodedenglu] = self.dataList['wodedenglu']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'我的退出':
                    ws['C4'] = OUTLOOPNUM
                    ws['C5'] = self.dataList['wodetuichu']['ANR'] + self.dataList['wodetuichu']['JRTERROR'] + self.dataList['wodetuichu']['JRTCRASH'] + self.dataList['wodetuichu']['APPDIED'] + self.dataList['wodetuichu']['SYSTEMERROR']
                    ws['C6'] = self.dataList['wodetuichu']['ANR']
                    ws['C7'] = self.dataList['wodetuichu']['JRTERROR'] + self.dataList['wodetuichu']['JRTCRASH']
                    ws['C8'] = self.dataList['wodetuichu']['APPDIED'] + self.dataList['wodetuichu']['SYSTEMERROR']
                    if len(self.dataList['wodetuichu']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['wodetuichu']['ERRORLOG'])):
                            log_wodetuichu = 'A' + str(i + 11)
                            line_wodetuichu = 'E' + str(i + 11)
                            message_wodetuichu = 'F' + str(i + 11)
                            ws[line_wodetuichu] = self.dataList['wodetuichu']['ERRORNUM'][i]
                            ws[message_wodetuichu] = self.dataList['wodetuichu']['ERRORLOG'][i]
                            ws[log_wodetuichu] = self.dataList['wodetuichu']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
            xlsFile = os.path.join(self.reportPath, u'飞凡APP重点功能压力测试报告(%s).xlsx' % time.strftime("%Y%m%d"))
            wb.save(xlsFile)
            if os.path.exists(file):
                os.remove(file)
        except:
            print("no sheet in %s named %s" % file, sheet_name)


if __name__ == "__main__":
    handler = Handler('Android')
    resultsPath = "%s/report/stability/%s/%s" % ("/Users/uasd-qiaojx/Desktop", '2016/12/29', '10')
    if os.path.exists(resultsPath):
        handler.handle(resultsPath)
