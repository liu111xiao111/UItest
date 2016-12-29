'''
Created on Nov 17, 2016

@author: songbo
'''

import os
import time
import datetime
import argparse
import xlsxwriter
import shutil
import openpyxl
import xlrd
# import pylab as pl

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

from tools.utility.constants import *
from configs.androidConfig import appVersion, phoneVersion, buildVersion, deviceNet, mtuanAPPVersion

# pl.mpl.rcParams['font.sans-serif'] = ['SimHei']

class Parser(object):
    def __init__(self):
        self.xmlTypeDict = {
            TYPE_CPU         : 'cpuUsage',
            TYPE_MEMORY      : 'ramUsage',
            TYPE_TX          : 'netUpFlow',
            TYPE_RX          : 'netDownFlow',
            TYPE_TEMPERATURE : 'temperature',
            TYPE_TRAFFIC     : 'traffic'
        }

    def txtParser(self, txtPath, type):
        performanceData = list()
        dataFile = open(txtPath, mode='r', encoding='utf-8')
        try:
            allLines = dataFile.readlines()
            for line in allLines:
                performanceData.append(line)
        except Exception as e:
            raise FileExistsError(e)
        finally:
            dataFile.close()

        return self._getTxtData(performanceData, type)

    def _getTxtData(self, performanceData, type):
        dataList = list()
        try:
            if type == TYPE_FPS:
                for line in performanceData:
                    value = str(line).split(' ')
                    if(len(value) > 1):
                        dataList.append(value)
            else:
                for line in performanceData:
                    value = str(line).split(':')
                    if(len(value) > 1):
                        dataList.append(value[1])
        except Exception as e:
            print(str(e))
        finally:
            return dataList

    def xmlParser(self, xmlPath, type):
        doc = ET.parse(xmlPath)
        root = doc.getroot()

        nodes = root.findall('performance')

        dataList = []
        if type == TYPE_TRAFFIC:
            dataList_rx = []
            dataList_tx = []
            for child in nodes:
                data_rx = self._getXmlData(child, self.xmlTypeDict[TYPE_RX])
                data_tx = self._getXmlData(child, self.xmlTypeDict[TYPE_TX])
                dataList_rx.append(float(data_rx))
                dataList_tx.append(float(data_tx))
            dataList.append(round(sum(dataList_rx) + sum(dataList_tx), 2))
        else:
            for child in nodes:
                data = self._getXmlData(child, self.xmlTypeDict[type])
                dataList.append(data)
        return dataList

    def _getXmlData(self, node, type):
        data = node.find(type).text
        if type == self.xmlTypeDict[TYPE_MEMORY]:
            data = round(float(data) / 1024, 2)
        return data if data else 0


class DataHandler(object):
    def __init__(self):
        self.rsPath = ''
        self.testCase = ''
        self.type = ''
        self.parser = Parser()

    def handle(self, rsPath, testCase, type):
        self.rsPath = rsPath
        self.testCase = testCase
        self.type = type

        return self._getPerfData()

    def _getPerfData(self):
        ffanDate = []
        mtuanDate = []
        if self.type == TYPE_TRAFFIC and (self.testCase == CASE_COLD_BOOT or self.testCase == CASE_WARM_BOOT):
            FFanTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, FFAN), CASE_FOLDER_LIST[self.testCase]),
                                       TRAFFIC_FILE)
            MTUANTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, MTUAN), CASE_FOLDER_LIST[self.testCase]),
                                        TRAFFIC_FILE)
            if os.path.exists(FFanTxtPath) and os.path.exists(MTUANTxtPath):
                ffanDate = self.parser.txtParser(FFanTxtPath, self.type)
                mtuanDate = self.parser.txtParser(MTUANTxtPath, self.type)
            else:
                print('Can not find %s test performance results!!!' % self.testCase)
        else:
            if self.type == TYPE_COLD_BOOT or self.type == TYPE_WARM_BOOT:
                FFanTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, FFAN), CASE_FOLDER_LIST[self.testCase]),
                                           BOOTTIME_FILE)
                MTUANTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, MTUAN), CASE_FOLDER_LIST[self.testCase]),
                                           BOOTTIME_FILE)
                if os.path.exists(FFanTxtPath) and os.path.exists(MTUANTxtPath):
                    ffanDate = self.parser.txtParser(FFanTxtPath, self.type)
                    mtuanDate = self.parser.txtParser(MTUANTxtPath, self.type)
                else:
                    print('Can not find %s test performance results!!!' % self.testCase)
            elif self.type == TYPE_FPS:
                FFanTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, FFAN), CASE_FOLDER_LIST[self.testCase]),
                                           FPS_FILE)
                MTuanTxtPath = os.path.join(os.path.join(os.path.join(self.rsPath, MTUAN), CASE_FOLDER_LIST[self.testCase]),
                                           FPS_FILE)
                if os.path.exists(FFanTxtPath) and os.path.exists(MTuanTxtPath):
                    ffanDate = self.parser.txtParser(FFanTxtPath, self.type)
                    mtuanDate = self.parser.txtParser(MTuanTxtPath, self.type)
                else:
                    print('Can not find %s test performance results!!!' % self.testCase)
            else:
                FFanXmlPath = os.path.join(os.path.join(os.path.join(self.rsPath, FFAN), CASE_FOLDER_LIST[self.testCase]), PERF_FILE)
                MTuanXmlPath = os.path.join(os.path.join(os.path.join(self.rsPath, MTUAN), CASE_FOLDER_LIST[self.testCase]), PERF_FILE)
                if os.path.exists(FFanXmlPath) and os.path.exists(MTuanXmlPath):
                    ffanDate = self.parser.xmlParser(FFanXmlPath, self.type)
                    mtuanDate = self.parser.xmlParser(MTuanXmlPath, self.type)
                else:
                    print('Can not find %s test performance results!!!' % self.testCase)

        return ffanDate, mtuanDate


# 性能数据处理装饰器
def dataHandle(type):
    def handler(func):
        def wrapper(cls, testCase):
            print('Process %s performance data!!!' % type)
            dataHandler = DataHandler()
            cls.dataList[FFAN], cls.dataList[MTUAN] = dataHandler.handle(rsPath=cls.rsPath,
                                                                         testCase=testCase,
                                                                         type=type)
            if cls.dataList[FFAN] and cls.dataList[MTUAN]:
                cls.dataLength = cls._dataLength(cls.dataList[FFAN], cls.dataList[MTUAN])
                func(cls, testCase)
        return wrapper
    return handler


class Handler(object):
    def __init__(self):
        self.rsPath = ''
        self.reportPath = ''
        self.workbook = ''
        self.dataLength = 0

        # 性能测试结果list
        self.dataList = dict()

        self.dataList[FFAN]  = ''
        self.dataList[MTUAN] = ''

    def handle(self, rsPath):
        self.rsPath = rsPath
        self._mkReportDir()
        try:
            for testCase in CASE_LIST:
                if os.path.exists(os.path.join(os.path.join(self.rsPath, FFAN), CASE_FOLDER_LIST[testCase])) and \
                        os.path.exists(os.path.join(os.path.join(self.rsPath, MTUAN), CASE_FOLDER_LIST[testCase])):
                    report = os.path.join(self.reportPath, EXCEL_REPORT_FILE % CASE_FOLDER_LIST[testCase])
                    self.workbook = xlsxwriter.Workbook(report)
                    if testCase == CASE_FPS:
                        self._fpsHandle(testCase)
                    else:
                        self._trafficHandle(testCase)
                        if testCase == CASE_WARM_BOOT:
                            self._warmBootHandle(testCase)
                        elif testCase == CASE_COLD_BOOT:
                            self._coldBootHandle(testCase)
                        else:
                            self._cpuHandle(testCase)
                            self._memoryHandle(testCase)
                            self._rxHandle(testCase)
                            self._txHandle(testCase)
                            self._temperatureHandle(testCase)
                else:
                    print('Missing %s test cases in result path.' % testCase)
        except Exception as e:
            print(e)
        finally:
            self.workbook.close()

    def _mkReportDir(self):
        '''
        创建报告目录, 包含excel报告, 性能测试结果图表图片
        '''
        self.reportPath = os.path.join(self.rsPath, 'report')
        if os.path.exists(self.reportPath):
            timestamp = os.path.getmtime(self.reportPath)
            date = datetime.datetime.fromtimestamp(timestamp).strftime('%Y_%m_%d_%H_%M_%S')
            newDirName = 'report_' + date

            newLogDir = os.path.join(self.rsPath, newDirName)
            try:
                os.renames(self.reportPath, newLogDir)
            except Exception as e:
                raise IOError('Modify the [%s] directory name failed with following error: \n'
                              '%s' % (self.reportPath, e))

        try:
            os.makedirs(self.reportPath)
        except Exception as e:
            raise IOError('Create the [%s] directory failed with following error: \n'
                          '%s' % (self.reportPath, e))

    @staticmethod
    def _dataLength(x, y):
        return len(x) if len(x) > len(y) else len(y)

    @dataHandle(TYPE_CPU)
    def _cpuHandle(self, testCase):
        # 生成CPU perf sheet
        self._createExcelReport(u'CPU 性能', self.workbook, TYPE_CPU, u'次数', u'cpu使用率(%)')

    @dataHandle(TYPE_MEMORY)
    def _memoryHandle(self, testCase):
        # 生成memory perf sheet
        self._createExcelReport(u'内存性能', self.workbook, TYPE_MEMORY, u'次数', u'内存(Mb)')

    @dataHandle(TYPE_RX)
    def _rxHandle(self, testCase):
        # 生成rx perf sheet
        self._createExcelReport(u'下行速率', self.workbook, TYPE_RX, u'次数', u'下行速率(KBps)')

    @dataHandle(TYPE_TX)
    def _txHandle(self, testCase):
        # 生成rx perf sheet
        self._createExcelReport(u'上行速率', self.workbook, TYPE_TX, u'次数', u'上行速率(KBps)')

    @dataHandle(TYPE_TEMPERATURE)
    def _temperatureHandle(self, testCase):
        # 生成rx perf sheet
        self._createExcelReport(u'电池温度', self.workbook, TYPE_TEMPERATURE, u'次数', u'电池温度(℃)')

    @dataHandle(TYPE_COLD_BOOT)
    def _coldBootHandle(self, testCase):
        # 生成冷启动性能excel
        self._createExcelReport(u'冷启动性能', self.workbook, TYPE_COLD_BOOT, u'次数', u'启动时间(ms)')

    @dataHandle(TYPE_WARM_BOOT)
    def _warmBootHandle(self, testCase):
        # 生成热启动性能excel
        self._createExcelReport(u'热启动性能', self.workbook, TYPE_WARM_BOOT, u'次数', u'启动时间(ms)')

    @dataHandle(TYPE_FPS)
    def _fpsHandle(self, testCase):
        self._createExcelReport(u'OverDraw和FPS 性能', self.workbook, TYPE_FPS, u'帧/秒', 'OverDraw', 'FPS')

    @dataHandle(TYPE_TRAFFIC)
    def _trafficHandle(self, testCase):
        self._createExcelReport(u'流量统计', self.workbook, TYPE_TRAFFIC, u'', u'流量统计(Kb)')

    def _createExcelReport(self, title, workbook, key, *args):
        worksheet = workbook.add_worksheet(title)

        format_title = workbook.add_format()  # 定义format_title格式对象
        format_title.set_border(1)  # 定义format_title对象单元格边框加粗(1像素)的格式
        format_title.set_bg_color('#cccccc')  # 定义format_title对象单元格背景颜色为
        # '#cccccc'的格式
        format_title.set_align('center')  # 定义format_title对象单元格居中对齐的格式
        format_title.set_bold()  # 定义format_title对象单元格内容加粗的格式

        format_ave = workbook.add_format()  # 定义format_ave格式对象
        format_ave.set_border(1)  # 定义format_ave对象单元格边框加粗(1像素)的格式

        if len(args) == 2 and key == TYPE_TRAFFIC:
            worksheet.write(1, 1, args[1], format_title)
            worksheet.write(2, 1, FFAN_APP, format_title)
            worksheet.write(3, 1, MTUAN_APP, format_title)
            worksheet.write(1, 2, '', format_title)
            worksheet.write(2, 2, float(self.dataList[FFAN][0]), format_ave)
            worksheet.write(3, 2, float(self.dataList[MTUAN][0]), format_ave)
            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({'values': [title, 2, 2, 2, 2],
                              'name': [title, 2, 1]})
            chart.add_series({'values': [title, 3, 2, 3, 2],
                              'name': [title, 3, 1]})
            chart.set_title({'name': title})
            chart.set_y_axis({'name': args[1]})
            worksheet.insert_chart('F15', chart, {'x_scale': 2, 'y_scale': 1})
        elif len(args) == 2 and (key == TYPE_COLD_BOOT or key == TYPE_WARM_BOOT):
            dataFFanList = [float(data) for data in self.dataList[FFAN]]
            dataMTuanList = [float(data) for data in self.dataList[MTUAN]]
            maxFFanData = max(dataFFanList)
            maxMTuanData = max(dataMTuanList)
            minFFanData = min(dataFFanList)
            minMTuanData = min(dataMTuanList)
            averageFFanData = round(float(sum(dataFFanList) / len(dataFFanList)), 2)
            averageMTuanData = round(float(sum(dataMTuanList) / len(dataMTuanList)), 2)
            worksheet.write(14, 1, u'统计', format_title)
            worksheet.write(14, 2, FFAN_APP, format_title)
            worksheet.write(14, 3, MTUAN_APP, format_title)
            worksheet.write(15, 1, u'最大值', format_title)
            worksheet.write(16, 1, u'最小值', format_title)
            worksheet.write(17, 1, u'平均值', format_title)
            worksheet.write(15, 2, maxFFanData, format_ave)
            worksheet.write(16, 2, minFFanData, format_ave)
            worksheet.write(17, 2, averageFFanData, format_ave)
            worksheet.write(15, 3, maxMTuanData, format_ave)
            worksheet.write(16, 3, minMTuanData, format_ave)
            worksheet.write(17, 3, averageMTuanData, format_ave)

            if len(self.dataList[FFAN]) > len(self.dataList[MTUAN]):
                for _ in range(0, len(self.dataList[FFAN]) - len(self.dataList[MTUAN])):
                    self.dataList[MTUAN].append('0')
            else:
                for _ in range(0, len(self.dataList[MTUAN]) - len(self.dataList[FFAN])):
                    self.dataList[FFAN].append('0')

            worksheet.write(1, 1, args[1], format_title)
            worksheet.write(2, 1, FFAN_APP, format_title)
            worksheet.write(3, 1, MTUAN_APP, format_title)
            for row in range(0, self.dataLength):
                worksheet.write(1, row + 2, row+1, format_title)
                worksheet.write(2, row + 2, float(self.dataList[FFAN][row]), format_ave)
                worksheet.write(3, row + 2, float(self.dataList[MTUAN][row]), format_ave)



            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({'categories': [title, 1, 2, 1, row+2],
                              'values': [title, 2, 2, 2, row+2],
                              'name': [title, 2, 1]})
            chart.add_series({'categories': [title, 1, 2, 1, row+2],
                              'values': [title, 3, 2, 3, row+2],
                              'name': [title, 3, 1]})
            chart.set_title({'name': title})
            chart.set_y_axis({'name': args[1]})
            worksheet.insert_chart('F15', chart, {'x_scale': 2, 'y_scale': 1.5})
        elif len(args) == 2:
            dataFFanList = [float(data) for data in self.dataList[FFAN]]
            dataMTuanList = [float(data) for data in self.dataList[MTUAN]]
            maxFFanData = max(dataFFanList)
            maxMTuanData = max(dataMTuanList)
            minFFanData = min(dataFFanList)
            minMTuanData = min(dataMTuanList)
            averageFFanData = round(float(sum(dataFFanList) / len(dataFFanList)), 2)
            averageMTuanData = round(float(sum(dataMTuanList) / len(dataMTuanList)), 2)
            worksheet.write(14, 1, u'统计', format_title)
            worksheet.write(14, 2, FFAN_APP, format_title)
            worksheet.write(14, 3, MTUAN_APP, format_title)
            worksheet.write(15, 1, u'最大值', format_title)
            worksheet.write(16, 1, u'最小值', format_title)
            worksheet.write(17, 1, u'平均值', format_title)
            worksheet.write(15, 2, maxFFanData, format_ave)
            worksheet.write(16, 2, minFFanData, format_ave)
            worksheet.write(17, 2, averageFFanData, format_ave)
            worksheet.write(15, 3, maxMTuanData, format_ave)
            worksheet.write(16, 3, minMTuanData, format_ave)
            worksheet.write(17, 3, averageMTuanData, format_ave)

            if len(self.dataList[FFAN]) > len(self.dataList[MTUAN]):
                for _ in range(0, len(self.dataList[FFAN]) - len(self.dataList[MTUAN])):
                    self.dataList[MTUAN].append('0')
            else:
                for _ in range(0, len(self.dataList[MTUAN]) - len(self.dataList[FFAN])):
                    self.dataList[FFAN].append('0')

            worksheet.write(1, 1, args[1], format_title)
            worksheet.write(2, 1, FFAN_APP, format_title)
            worksheet.write(3, 1, MTUAN_APP, format_title)
            for row in range(0, self.dataLength):
                worksheet.write(1, row + 2, row + 1, format_title)
                worksheet.write(2, row + 2, float(self.dataList[FFAN][row]), format_ave)
                worksheet.write(3, row + 2, float(self.dataList[MTUAN][row]), format_ave)

            chart = workbook.add_chart({'type': 'line'})
            chart.add_series({'categories': [title, 1, 2, 1, row + 2],
                              'values': [title, 2, 2, 2, row + 2],
                              'name': [title, 2, 1]})
            chart.add_series({'categories': [title, 1, 2, 1, row + 2],
                              'values': [title, 3, 2, 3, row + 2],
                              'name': [title, 3, 1]})
            chart.set_title({'name': title})
            chart.set_x_axis({'name' : args[0]})
            chart.set_y_axis({'name': args[1]})
            worksheet.insert_chart('F15', chart, {'x_scale': 3.5, 'y_scale': 1.5})
        else:
            worksheet.write(1, 1, args[1], format_title)
            worksheet.write(2, 1, FFAN_APP, format_title)
            worksheet.write(3, 1, MTUAN_APP, format_title)
            worksheet.write(1, 7, args[2], format_title)
            worksheet.write(2, 7, FFAN_APP, format_title)
            worksheet.write(3, 7, MTUAN_APP, format_title)
            for data in self.dataList[FFAN]:
                if data[0] == 'Mine':
                    worksheet.write(1, 2, data[0], format_title)
                    worksheet.write(1, 8, data[0], format_title)
                    worksheet.write(2, 2, float(data[1]), format_ave)
                    worksheet.write(2, 8, float(data[2]), format_ave)
                elif data[0] == 'Dashboard':
                    worksheet.write(1, 3, data[0], format_title)
                    worksheet.write(1, 9, data[0], format_title)
                    worksheet.write(2, 3, float(data[1]), format_ave)
                    worksheet.write(2, 9, float(data[2]), format_ave)
                elif data[0] == 'BenefitsLife':
                    worksheet.write(1, 4, data[0], format_title)
                    worksheet.write(1, 10, data[0], format_title)
                    worksheet.write(2, 4, float(data[1]), format_ave)
                    worksheet.write(2, 10, float(data[2]), format_ave)
                else:
                    worksheet.write(1, 5, data[0], format_title)
                    worksheet.write(1, 11, data[0], format_title)
                    worksheet.write(2, 5, float(data[1]), format_ave)
                    worksheet.write(2, 11, float(data[2]), format_ave)
            for data in self.dataList[MTUAN]:
                if data[0] == 'Mine':
                    worksheet.write(3, 2, float(data[1]), format_ave)
                    worksheet.write(3, 8, float(data[2]), format_ave)
                elif data[0] == 'Dashboard':
                    worksheet.write(3, 3, float(data[1]), format_ave)
                    worksheet.write(3, 9, float(data[2]), format_ave)
            worksheet.write(3, 4, 0, format_ave)
            worksheet.write(3, 5, 0, format_ave)
            worksheet.write(3, 10, 0, format_ave)
            worksheet.write(3, 11, 0, format_ave)

            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({'categories': [title, 1, 2, 1, 5],
                              'values': [title, 2, 2, 2, 5],
                              'name': [title, 2, 1]})
            chart.add_series({'categories': [title, 1, 2, 1, 5],
                              'values': [title, 3, 2, 3, 5],
                              'name': [title, 3, 1]})
            chart.set_title({'name': 'OverDraw 性能'})
            chart.set_y_axis({'name': args[0]})
            worksheet.insert_chart('B14', chart, {'x_scale': 1.5, 'y_scale': 1})
            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({'categories': [title, 1, 8, 1, 11],
                              'values': [title, 2, 8, 2, 11],
                              'name': [title, 2, 7]})
            chart.add_series({'categories': [title, 1, 8, 1, 11],
                              'values': [title, 3, 8, 3, 11],
                              'name': [title, 3, 7]})
            chart.set_title({'name': 'FPS 性能'})
            chart.set_y_axis({'name': args[0]})
            worksheet.insert_chart('N14', chart, {'x_scale': 1.5, 'y_scale': 1})

class DataSummary(object):
    def __init__(self):
        self.rsPath = ''
        self.testCase = ''

    def handle(self, rsPath):
        self.rsPath = rsPath
        self.attachmentPath = os.path.join(self.rsPath, 'attachment')
        if not os.path.exists(self.attachmentPath):
            os.makedirs(self.attachmentPath)

    def getLogData(self,testCase):
        self.testCase = testCase
        return self._parserLogData()

    def _parserLogData(self):
        i = 0
        caseErrorInfo = {'ANR': 0, 'JRTERROR': 0, 'JRTCRASH': 0, 'APPDIED': 0, 'SYSTEMERROR': 0}
        tmpFile = os.path.join(self.logPath, '%s.txt') % self.testCase
        cmdFind = 'find %s -name "%s*.log" > %s' % (self.logPath, self.testCase, tmpFile)
        os.system(cmdFind)
        if os.path.exists(tmpFile):
            logCasePath = open(tmpFile, 'r')
            logPaths = logCasePath.readlines()
            logCasePath.close()
            if logPaths != []:
                for logPath in logPaths:
                    caseLogFile = logPath[:-1]
                    if os.path.exists(caseLogFile):
                        logInfo = open(caseLogFile, 'r')
                        logLines = logInfo.readlines()
                        logInfo.close()
                        for logLine in logLines:
                            if logLine.find("anr") != -1:
                                caseErrorInfo['ANR'] += 1
                            elif logLine.find("crash") != -1:
                                caseErrorInfo['JRTCRASH'] += 1
                            elif logLine.find("Reading a NULL string not supported here.") != -1:
                                i += 1
                                caseErrorInfo['JRTERROR'] += 1
                            elif logLine.find("Got null root node from accessibility - Retrying...") != -1:
                                caseErrorInfo['JRTERROR'] += 1
                            elif logLine.find("died") != -1:
                                caseErrorInfo['APPDIED'] += 1
                            elif logLine.find("system error") != -1:
                                caseErrorInfo['SYSTEMERROR'] += 1
                caseErrorInfo['JRTERROR'] = caseErrorInfo['JRTERROR'] - (i-1)

        if os.path.exists(tmpFile):
            os.remove(tmpFile)

        return caseErrorInfo

class PerformanceSummary(object):
    def __init__(self):
        self.rsPath = ''
        self.attachmentPath = ''
        self.workbook = ''
        self.cpuData = {'dianying': {'max':'100%', 'min':'0%', 'anv':'50%', 'top':u'第1s', '40%':u'1s', '60%':u'1s', '80%':u'1s', '100%':u'1s', 'rst':u'正常'}, 
                        'meishi': {'max':'100%', 'min':'0%', 'anv':'50%', 'top':u'第1s', '40%':u'1s', '60%':u'1s', '80%':u'1s', '100%':u'1s', 'rst':u'正常'},
                        'dingdan': {'max':'100%', 'min':'0%', 'anv':'50%', 'top':u'第1s', '40%':u'1s', '60%':u'1s', '80%':u'1s', '100%':u'1s', 'rst':u'正常'},
                        'denglu': {'max':'100%', 'min':'0%', 'anv':'50%', 'top':u'第1s', '40%':u'1s', '60%':u'1s', '80%':u'1s', '100%':u'1s', 'rst':u'正常'}}
        self.memoryData = {'dianying': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60M':u'1s', '80M':u'1s', '100M':u'1s', 'rst':u'正常'},
                           'meishi': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60M':u'1s', '80M':u'1s', '100M':u'1s', 'rst':u'正常'},
                           'dingdan': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60M':u'1s', '80M':u'1s', '100M':u'1s', 'rst':u'正常'},
                           'denglu': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60M':u'1s', '80M':u'1s', '100M':u'1s', 'rst':u'正常'}}
        self.txData = {'dianying': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '10K':u'1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'meishi': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '10K':u'1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'dingdan': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '10K':u'1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'denglu': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '10K':u'1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'}}
        self.rxData = {'dianying': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'meishi': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'dingdan': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'},
                       'denglu': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '20K':u'1s', '40K':u'1s', '60K':u'1s', '80K':u'1s', '100K':u'1s','rst':u'正常'}}
        self.batteryTemperatureData = {'dianying': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60℃':u'1s', '80℃':u'1s', 'rst':u'正常'},
                            'meishi': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60℃':u'1s', '80℃':u'1s', 'rst':u'正常'},
                            'dingdan': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60℃':u'1s', '80℃':u'1s', 'rst':u'正常'},
                            'denglu': {'max':'100', 'min':'0', 'anv':'50', 'top':u'第1s', '60℃':u'1s', '80℃':u'1s', 'rst':u'正常'}}
        self.fpsData = {'max':'10', 'min':'0', 'anv':'50', '16ms':u'否', '20ms':u'否', 'rst':u'正常'}
        self.coldBootData = {'num':'10', 'max':'9999', 'min':'1000', 'anv':u'5555', 'top':u'第1次', 'rst':u'正常'}
        self.warmBootData = {'num':'10', 'max':'1000', 'min':'0', 'anv':u'500', 'top':u'第1次', 'rst':u'正常'}

        self.ffanAppVersion = appVersion
        self.mtuanAppVersion = mtuanAPPVersion
        self.phoneVersion = phoneVersion
        self.buildVersion = buildVersion
        self.deviceNet = deviceNet

    def performanceSummary(self, rsPath):
        self.caseData = {u'CPU 性能':['0'], u'内存性能':['0'], u'上行速率':['0'], u'下行速率':['0'], u'电池温度':['0']}
        self.rsPath = rsPath
        try:
            resourcesDirectory = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/resources/"
            xlsFileTemplate = os.path.join(resourcesDirectory, 'templatePerformance.xlsx')
            self.attachmentPath = os.path.join(self.rsPath, 'attachment')
            if not os.path.exists(self.attachmentPath):
                os.makedirs(self.attachmentPath)
            shutil.copy(xlsFileTemplate, self.attachmentPath)
            reportPath = os.path.join(self.rsPath, 'report')
            if os.path.exists(reportPath):
                dianYingFile = os.path.join(reportPath, 'dianying_performance_result.xlsx')
                meiShiHuiFile = os.path.join(reportPath, 'meishihui_performance_result.xlsx')
                dingDanFile = os.path.join(reportPath, 'wodedingdan_performance_result.xlsx')
                dengLuFile = os.path.join(reportPath, 'wodedenglu_performance_result.xlsx')
                fpsFile = os.path.join(reportPath, 'fps_performance_result.xlsx')
                coldBootFile = os.path.join(reportPath, 'coldboot_performance_result.xlsx')
                warmBootFile = os.path.join(reportPath, 'warmboot_performance_result.xlsx')
                excelList = {'dianying':dianYingFile, 'meishi':meiShiHuiFile, 'dingdan':dingDanFile, 'denglu':dengLuFile, 'fps':fpsFile, 'coldboot':coldBootFile, 'warmboot':warmBootFile}
                for (case,excelFile) in excelList.items():
                    if os.path.exists(excelFile):
                        print(excelFile)
                        data = xlrd.open_workbook(excelFile)
                        if excelFile in (dianYingFile, meiShiHuiFile, dingDanFile, dengLuFile):
                            sheetList = [u'CPU 性能', u'内存性能', u'上行速率', u'下行速率', u'电池温度']
                            caseTime = 0
                            for sheetTemp in sheetList:
                                table = data.sheet_by_name(sheetTemp)
                                if sheetTemp == u'CPU 性能':
                                    print("XX")
                                    self.cpuData[case]['max'] = table.cell(15,2).value
                                    self.cpuData[case]['min'] = table.cell(16,2).value
                                    self.cpuData[case]['anv'] = table.cell(17,2).value
                                    cpuOver40 = 0
                                    cpuOver60 = 0
                                    cpuOver80 = 0
                                    cpuOver100 = 0
                                    cpuTop = []
                                    cpuTempData = table.row_values(2)
                                    for q in range(2, len(cpuTempData)):
                                        if (40 < cpuTempData[q] and cpuTempData[q] < 60) or (cpuTempData[q] == 60):
                                            cpuOver40 += 1
                                        elif (60 < cpuTempData[q] and cpuTempData[q] < 80) or (cpuTempData[q] == 80):
                                            cpuOver60 += 1
                                        elif (80 < cpuTempData[q] and cpuTempData[q] < 100) or (cpuTempData[q] == 100):
                                            cpuOver80 += 1
                                        elif 100 < cpuTempData[q]:
                                            cpuOver100 += 1
                                        if cpuTempData[q] == self.cpuData[case]['max']:
                                            cpuTop.append(q-1)
                                    self.cpuData[case]['40%'] = str(cpuOver40) + 's'
                                    self.cpuData[case]['60%'] = str(cpuOver60) + 's'
                                    self.cpuData[case]['80%'] = str(cpuOver80) + 's'
                                    self.cpuData[case]['100%'] = str(cpuOver100) + 's'
                                    if len(cpuTop) > 1:
                                        cpuTopTotal = ''
                                        for r in range(len(cpuTop)):
                                            if r == 0:
                                                cpuTopTotal = str(cpuTop[r])
                                            else:
                                                cpuTopTotal = cpuTopTotal + '、' + str(cpuTop[r])
                                            self.cpuData[case]['top'] = u'第' + cpuTopTotal + 's'
                                    else:
                                        self.cpuData[case]['top'] = u'第' + str(cpuTop[0]) + 's'
                                    cpuTotalTime = len(cpuTempData) - 2
                                    for dd in range(len(cpuTempData), 2, -1):
                                        if cpuTempData[dd-1] == 0:
                                            cpuTotalTime -= 1
                                    if (cpuOver60 > cpuTotalTime * 0.4) and ((cpuOver60 < cpuTotalTime * 0.6) or (cpuOver60 == cpuTotalTime * 0.6)):
                                        self.cpuData[case]['rst'] = u'偏高'
                                    elif cpuOver60 > cpuTotalTime * 0.6:
                                        self.cpuData[case]['rst'] = u'高'
                                    else:
                                        self.cpuData[case]['rst'] = u'正常'
                                    caseTime = cpuTotalTime
                                elif sheetTemp == u'内存性能':
                                    print("XX")
                                    self.memoryData[case]['max'] = table.cell(15,2).value
                                    self.memoryData[case]['min'] = table.cell(16,2).value
                                    self.memoryData[case]['anv'] = table.cell(17,2).value
                                    memoryOver60 = 0
                                    memoryOver80 = 0
                                    memoryOver100 = 0
                                    memoryOver120 = 0
                                    memoryTop = []
                                    memoryTempData = table.row_values(2)
                                    for t in range(2, len(memoryTempData)):
                                        if (60 < memoryTempData[t] and memoryTempData[t] < 80) or (memoryTempData[t] == 80):
                                            memoryOver60 += 1
                                        elif (80 < memoryTempData[t] and memoryTempData[t] < 100) or (memoryTempData[t] == 100):
                                            memoryOver60 += 1
                                            memoryOver80 += 1
                                        elif (100 < memoryTempData[t] and memoryTempData[t] < 100) or (memoryTempData[t] == 120):
                                            memoryOver60 += 1
                                            memoryOver80 += 1
                                            memoryOver100 += 1
                                        elif 120 < memoryTempData[t]:
                                            memoryOver60 += 1
                                            memoryOver80 += 1
                                            memoryOver100 += 1
                                            memoryOver120 += 1
                                        if memoryTempData[t] == self.memoryData[case]['max']:
                                            memoryTop.append(t-1)
                                    self.memoryData[case]['60M'] = str(memoryOver60) + 's'
                                    self.memoryData[case]['80M'] = str(memoryOver80) + 's'
                                    self.memoryData[case]['100M'] = str(memoryOver100) + 's'
                                    if len(memoryTop) > 1:
                                        memoryTopTotal = ''
                                        for u in range(len(memoryTop)):
                                            if u == 0:
                                                memoryTopTotal = str(memoryTop[u])
                                            else:
                                                memoryTopTotal = memoryTopTotal + '、' + str(memoryTop[u])
                                            self.memoryData[case]['top'] = u'第' + memoryTopTotal + u's'
                                    else:
                                        self.memoryData[case]['top'] = u'第' + str(memoryTop[0]) + u's'
                                    memoryTotalTime = len(memoryTempData) - 2
                                    for cc in range(len(memoryTempData), 2, -1):
                                        if memoryTempData[cc-1] == 0:
                                            memoryTotalTime -= 1
                                    if (memoryOver80 > memoryTotalTime * 0.5) and ((memoryOver120 < memoryTotalTime * 0.5) or (memoryOver120 == memoryTotalTime * 0.5)):
                                        self.memoryData[case]['rst'] = u'偏高'
                                    elif memoryOver120 > memoryTotalTime * 0.5:
                                        self.memoryData[case]['rst'] = u'高'
                                    else:
                                        self.memoryData[case]['rst'] = u'正常'
                                elif sheetTemp == u'上行速率':
                                    self.txData[case]['max'] = table.cell(15,2).value
                                    self.txData[case]['min'] = table.cell(16,2).value
                                    self.txData[case]['anv'] = table.cell(17,2).value
                                    txOver10 = 0
                                    txOver20 = 0
                                    txOver40 = 0
                                    txOver60 = 0
                                    txOver80 = 0
                                    txOver100 = 0
                                    txTop = []
                                    txTempData = table.row_values(2)
                                    for v in range(2, len(txTempData)):
                                        if (10 < txTempData[v] and txTempData[v] < 20) or (txTempData[v] == 20):
                                            txOver10 += 1
                                        elif (20 < txTempData[v] and txTempData[v] < 40) or (txTempData[v] == 40):
                                            txOver10 += 1
                                            txOver20 += 1
                                        elif (40 < txTempData[v] and txTempData[v] < 60) or (txTempData[v] == 60):
                                            txOver10 += 1
                                            txOver20 += 1
                                            txOver40 += 1
                                        elif (60 < txTempData[v] and txTempData[v] < 80) or (txTempData[v] == 80):
                                            txOver10 += 1
                                            txOver20 += 1
                                            txOver40 += 1
                                            txOver60 += 1
                                        elif (80 < txTempData[v] and txTempData[v] < 100) or (txTempData[v] == 100):
                                            txOver10 += 1
                                            txOver20 += 1
                                            txOver40 += 1
                                            txOver60 += 1
                                            txOver80 += 1
                                        elif 100 < txTempData[v]:
                                            txOver10 += 1
                                            txOver20 += 1
                                            txOver40 += 1
                                            txOver60 += 1
                                            txOver80 += 1
                                            txOver100 += 1
                                        if txTempData[v] == self.txData[case]['max']:
                                            txTop.append(v-1)
                                    self.txData[case]['10K'] = str(txOver10) + 's'
                                    self.txData[case]['20K'] = str(txOver20) + 's'
                                    self.txData[case]['40K'] = str(txOver40) + 's'
                                    self.txData[case]['60K'] = str(txOver60) + 's'
                                    self.txData[case]['80K'] = str(txOver80) + 's'
                                    self.txData[case]['100K'] = str(txOver100) + 's'
                                    if len(txTop) > 1:
                                        txTopTotal = ''
                                        for w in range(len(txTop)):
                                            if w == 0:
                                                txTopTotal = str(txTop[w])
                                            else:
                                                txTopTotal = txTopTotal + '、' + str(txTop[w])
                                            self.txData[case]['top'] = u'第' + txTopTotal + u's'
                                    else:
                                        self.txData[case]['top'] = u'第' + str(txTop[0]) + u's'
                                    if (txOver40 > caseTime * 0.5) and ((txOver60 < caseTime * 0.5) or (txOver60 == caseTime * 0.5)):
                                        self.txData[case]['rst'] = u'偏高'
                                    elif txOver60 > caseTime * 0.5:
                                        self.txData[case]['rst'] = u'高'
                                    else:
                                        self.txData[case]['rst'] = u'正常'
                                elif sheetTemp == u'下行速率':
                                    self.rxData[case]['max'] = table.cell(15,2).value
                                    self.rxData[case]['min'] = table.cell(16,2).value
                                    self.rxData[case]['anv'] = table.cell(17,2).value
                                    rxOver20 = 0
                                    rxOver30 = 0
                                    rxOver40 = 0
                                    rxOver60 = 0
                                    rxOver80 = 0
                                    rxOver100 = 0
                                    rxTop = []
                                    rxTempData = table.row_values(2)
                                    for x in range(2, len(rxTempData)):
                                        if (20 < rxTempData[x] and rxTempData[x] < 30) or (rxTempData[x] == 30):
                                            rxOver20 += 1
                                        elif (30 < rxTempData[x] and rxTempData[x] < 40) or (rxTempData[x] == 40):
                                            rxOver20 += 1
                                            rxOver30 += 1
                                        elif (40 < rxTempData[x] and rxTempData[x] < 60) or (rxTempData[x] == 60):
                                            rxOver20 += 1
                                            rxOver40 += 1
                                        elif (60 < rxTempData[x] and rxTempData[x] < 80) or (rxTempData[x] == 80):
                                            rxOver20 += 1
                                            rxOver40 += 1
                                            rxOver60 += 1
                                        elif (80 < rxTempData[x] and rxTempData[x] < 100) or (rxTempData[x] == 100):
                                            rxOver20 += 1
                                            rxOver40 += 1
                                            rxOver60 += 1
                                            rxOver80 += 1
                                        elif 100 < rxTempData[x]:
                                            rxOver20 += 1
                                            rxOver40 += 1
                                            rxOver60 += 1
                                            rxOver80 += 1
                                            rxOver100 += 1
                                        if rxTempData[x] == self.rxData[case]['max']:
                                            rxTop.append(x-1)
                                    self.rxData[case]['20K'] = str(rxOver20) + 's'
                                    self.rxData[case]['40K'] = str(rxOver40) + 's'
                                    self.rxData[case]['60K'] = str(rxOver60) + 's'
                                    self.rxData[case]['80K'] = str(rxOver80) + 's'
                                    self.rxData[case]['100K'] = str(rxOver100) + 's'
                                    if len(rxTop) > 1:
                                        rxTopTotal = ''
                                        for y in range(len(rxTop)):
                                            if y == 0:
                                                rxTopTotal = str(rxTop[y])
                                            else:
                                                rxTopTotal = rxTopTotal + '、' + str(rxTop[y])
                                            self.rxData[case]['top'] = u'第' + rxTopTotal + u's'
                                    else:
                                        self.rxData[case]['top'] = u'第' + str(rxTop[0]) + u's'
                                    if (rxOver20 > caseTime * 0.4) and ((rxOver30 < caseTime * 0.4) or (rxOver30 == caseTime * 0.4)):
                                        self.rxData[case]['rst'] = u'偏高'
                                    elif rxOver30 > caseTime * 0.4:
                                        self.rxData[case]['rst'] = u'高'
                                    else:
                                        self.rxData[case]['rst'] = u'正常'
                                elif sheetTemp == u'电池温度':
                                    self.batteryTemperatureData[case]['max'] = table.cell(15,2).value
                                    self.batteryTemperatureData[case]['min'] = table.cell(16,2).value
                                    self.batteryTemperatureData[case]['anv'] = table.cell(17,2).value
                                    batteryTemperatureOver37 = 0
                                    batteryTemperatureOver38 = 0
                                    batteryTemperatureOver60 = 0
                                    batteryTemperatureOver80 = 0
                                    batteryTemperatureTop = []
                                    batteryTemperatureTempData = table.row_values(2)
                                    for aa in range(2, len(batteryTemperatureTempData)):
                                        if (60 < batteryTemperatureTempData[aa] and batteryTemperatureTempData[aa] < 80) or (batteryTemperatureTempData[aa] == 80):
                                            batteryTemperatureOver37 += 1
                                            batteryTemperatureOver38 += 1
                                            batteryTemperatureOver60 += 1
                                        elif 80 < batteryTemperatureTempData[aa]:
                                            batteryTemperatureOver37 += 1
                                            batteryTemperatureOver38 += 1
                                            batteryTemperatureOver60 += 1
                                            batteryTemperatureOver80 += 1
                                        if batteryTemperatureTempData[aa] == self.batteryTemperatureData[case]['max']:
                                            batteryTemperatureTop.append(aa-1)
                                    self.batteryTemperatureData[case]['60℃'] = str(batteryTemperatureOver60) + 's'
                                    self.batteryTemperatureData[case]['80℃'] = str(batteryTemperatureOver80) + 's'
                                    if len(batteryTemperatureTop) > 1:
                                        batteryTemperatureTopTotal = ''
                                        for bb in range(len(batteryTemperatureTop)):
                                            if bb == 0:
                                                batteryTemperatureTopTotal = str(batteryTemperatureTop[bb])
                                            elif len(batteryTemperatureTop) > 3:
                                                batteryTemperatureTopTotal = str(batteryTemperatureTop[0]) + '、' + str(batteryTemperatureTop[1]) + '...'
                                            else:
                                                batteryTemperatureTopTotal = batteryTemperatureTopTotal + '、' + str(batteryTemperatureTop[bb])
                                            self.batteryTemperatureData[case]['top'] = u'第' + batteryTemperatureTopTotal + u's'
                                    else:
                                        self.batteryTemperatureData[case]['top'] = u'第' + str(batteryTemperatureTop[0]) + u's'
                                    batteryTemperatureTotalTime = len(batteryTemperatureTempData) - 2
                                    for ee in range(len(batteryTemperatureTempData), 2, -1):
                                        if batteryTemperatureTempData[ee-1] == 0:
                                            batteryTemperatureTotalTime -= 1
                                    if (batteryTemperatureOver37 > batteryTemperatureTotalTime * 0.5) and ((batteryTemperatureOver38 < batteryTemperatureTotalTime * 0.5) or (batteryTemperatureOver38 == batteryTemperatureTotalTime * 0.5)):
                                        self.batteryTemperatureData[case]['rst'] = u'偏高'
                                    elif batteryTemperatureOver38 > batteryTemperatureTotalTime * 0.5:
                                        self.batteryTemperatureData[case]['rst'] = u'高'
                                    else:
                                        self.batteryTemperatureData[case]['rst'] = u'正常'
                        elif excelFile == fpsFile:
                            print("cc")
                            table = data.sheet_by_name(u'OverDraw和FPS 性能')
                            tempFpsDate = []
                            over16 = u'否'
                            over20 = u'否'
                            temp16 = {0: u'否', 1:u'是'}
                            temp20 = {0: u'否', 1:u'是'}
                            mine = table.cell(2,8).value
                            tempFpsDate.append(mine)
                            dashboard = table.cell(2,9).value
                            tempFpsDate.append(dashboard)
                            huilife = table.cell(2,10).value
                            tempFpsDate.append(huilife)
                            ffantong = table.cell(2,11).value
                            tempFpsDate.append(ffantong)
                            for i in range(len(tempFpsDate)):
                                if tempFpsDate[i] > 20:
                                    over16 = temp16[1]
                                    over20 = temp20[1]
                                elif (16 < tempFpsDate[i] < 20) or tempFpsDate[i] == 20:
                                    over16 = temp16[1]
                                    over20 = temp20[0]
                                elif over16 == temp16[1]:
                                    over16 = temp16[1]
                                elif over20 == temp20[1]:
                                    over20 = temp20[1]
                            self.fpsData['max'] = max(tempFpsDate)
                            self.fpsData['min'] = min(tempFpsDate)
                            self.fpsData['anv'] = round(float(sum(tempFpsDate) / len(tempFpsDate)), 2)
                            self.fpsData['16ms'] = over16
                            self.fpsData['20ms'] = over20
                            if over16 == u'是' and over20 == u'是':
                                self.fpsData['rst'] = u'高'
                            elif over16 == u'是' and over20 == u'否':
                                self.fpsData['rst'] = u'偏高'
                            else:
                                self.fpsData['rst'] = u'正常'
                        elif excelFile == coldBootFile:
                            if data.sheets()[1] == u'冷启动性能':
                                print("bb")
                                table = data.sheet_by_name(u'冷启动性能')
                                self.coldBootData['num'] = 10
                                self.coldBootData['max'] = table.cell(15,2).value
                                self.coldBootData['min'] = table.cell(16,2).value
                                self.coldBootData['anv'] = table.cell(17,2).value
                                coldTop = []
                                coldTempData = table.row_values(2)
                                coldMid = 0
                                coldBig = 0
                                for j in range(2, len(coldTempData)):
                                    if (coldTempData[j] > 4000 and coldTempData[j] < 6000) or (coldTempData[j] == 6000):
                                        coldMid += 1
                                    elif coldTempData[j] > 6000:
                                        coldBig += 1
                                    if coldTempData[j] == self.coldBootData['max']:
                                        coldTop.append(j-1)
                                if len(coldTop) > 1:
                                    for s in range(len(coldTop)):
                                        if s == 0:
                                            coldTopTotal = str(coldTop[s])
                                        else:
                                            coldTopTotal = coldTopTotal + '、' + str(coldTop[s])
                                        self.coldBootData['top'] = u'第' + coldTopTotal + u'次'
                                else:
                                    self.coldBootData['top'] = u'第' + str(coldTop[0]) + u'次'
                                if self.coldBootData['max'] > 4000 and coldMid > 3:
                                    self.coldBootData['rst'] = u'偏高'
                                elif self.coldBootData['max'] > 6000 and coldBig > 3:
                                    self.coldBootData['rst'] = u'高'
                                else:
                                    self.coldBootData['rst'] = u'正常'
                        elif excelFile == warmBootFile:
                            if data.sheets()[1] == u'热启动性能':
                                print("aa")
                                table = data.sheet_by_name(u'热启动性能')
                                self.warmBootData['num'] = 10
                                self.warmBootData['max'] = table.cell(15,2).value
                                self.warmBootData['min'] = table.cell(16,2).value
                                self.warmBootData['anv'] = table.cell(17,2).value
                                warmTop = []
                                warmTempData = table.row_values(2)
                                warmMid = 0
                                warmBig = 0
                                for k in range(2, len(warmTempData)):
                                    if (warmTempData[k] > 2000 and warmTempData[k] < 4000) or (warmTempData[k] == 4000):
                                        warmMid += 1
                                    elif warmTempData[k] > 4000:
                                        warmBig += 1
                                    if warmTempData[k] == self.warmBootData['max']:
                                        warmTop.append(k-1)
                                if len(warmTop) > 1:
                                    for n in range(len(warmTop)):
                                        if n == 0:
                                            warmTopTotal = str(warmTop[n])
                                        else:
                                            warmTopTotal = warmTopTotal + '、' + str(warmTop[n])
                                        self.warmBootData['top'] = u'第' + warmTopTotal + u'次'
                                else:
                                    self.warmBootData['top'] = u'第' + str(warmTop[0]) + u'次'
                                if self.warmBootData['max'] > 2000 and warmMid > 3:
                                    self.warmBootData['rst'] = u'偏高'
                                elif self.warmBootData['max'] > 4000 and warmBig > 3:
                                    self.warmBootData['rst'] = u'高'
                                else:
                                    self.warmBootData['rst'] = u'正常'
            xlsFile = os.path.join(self.attachmentPath, 'templatePerformance.xlsx')
        except Exception as e:
            print(e)
        finally:
            self._writeExcel(xlsFile)

    def _writeExcel(self, file='file.xls'):

        try:
            wb = openpyxl.load_workbook(file)
            for sheet_name in PERFORMANCE_REPORT_SHEET:
                ws = wb.get_sheet_by_name(sheet_name)
                if sheet_name == u'测试环境':
                    ws['C4'] = self.phoneVersion
                    ws['C5'] = self.buildVersion
                    ws['C9'] = self.ffanAppVersion
                    ws['C10'] = self.mtuanAppVersion
                    ws['A14'] = self.deviceNet
                elif sheet_name == u'CPU 性能':
                    ws['B3'] = self.cpuData['dianying']['max']
                    ws['C3'] = self.cpuData['dianying']['min']
                    ws['D3'] = self.cpuData['dianying']['anv']
                    ws['E3'] = self.cpuData['dianying']['top']
                    ws['F3'] = self.cpuData['dianying']['40%']
                    ws['G3'] = self.cpuData['dianying']['60%']
                    ws['H3'] = self.cpuData['dianying']['80%']
                    ws['I3'] = self.cpuData['dianying']['100%']
                    ws['J3'] = self.cpuData['dianying']['rst']
                    ws['B4'] = self.cpuData['meishi']['max']
                    ws['C4'] = self.cpuData['meishi']['min']
                    ws['D4'] = self.cpuData['meishi']['anv']
                    ws['E4'] = self.cpuData['meishi']['top']
                    ws['F4'] = self.cpuData['meishi']['40%']
                    ws['G4'] = self.cpuData['meishi']['60%']
                    ws['H4'] = self.cpuData['meishi']['80%']
                    ws['I4'] = self.cpuData['meishi']['100%']
                    ws['J4'] = self.cpuData['meishi']['rst']
                    ws['B5'] = self.cpuData['dingdan']['max']
                    ws['C5'] = self.cpuData['dingdan']['min']
                    ws['D5'] = self.cpuData['dingdan']['anv']
                    ws['E5'] = self.cpuData['dingdan']['top']
                    ws['F5'] = self.cpuData['dingdan']['40%']
                    ws['G5'] = self.cpuData['dingdan']['60%']
                    ws['H5'] = self.cpuData['dingdan']['80%']
                    ws['I5'] = self.cpuData['dingdan']['100%']
                    ws['J5'] = self.cpuData['dingdan']['rst']
                    ws['B6'] = self.cpuData['denglu']['max']
                    ws['C6'] = self.cpuData['denglu']['min']
                    ws['D6'] = self.cpuData['denglu']['anv']
                    ws['E6'] = self.cpuData['denglu']['top']
                    ws['F6'] = self.cpuData['denglu']['40%']
                    ws['G6'] = self.cpuData['denglu']['60%']
                    ws['H6'] = self.cpuData['denglu']['80%']
                    ws['I6'] = self.cpuData['denglu']['100%']
                    ws['J6'] = self.cpuData['denglu']['rst']
                elif sheet_name == u'内存性能':
                    ws['B3'] = self.memoryData['dianying']['max']
                    ws['C3'] = self.memoryData['dianying']['min']
                    ws['D3'] = self.memoryData['dianying']['anv']
                    ws['E3'] = self.memoryData['dianying']['top']
                    ws['F3'] = self.memoryData['dianying']['60M']
                    ws['G3'] = self.memoryData['dianying']['80M']
                    ws['H3'] = self.memoryData['dianying']['100M']
                    ws['I3'] = self.memoryData['dianying']['rst']
                    ws['B4'] = self.memoryData['meishi']['max']
                    ws['C4'] = self.memoryData['meishi']['min']
                    ws['D4'] = self.memoryData['meishi']['anv']
                    ws['E4'] = self.memoryData['meishi']['top']
                    ws['F4'] = self.memoryData['meishi']['60M']
                    ws['G4'] = self.memoryData['meishi']['80M']
                    ws['H4'] = self.memoryData['meishi']['100M']
                    ws['I4'] = self.memoryData['meishi']['rst']
                    ws['B5'] = self.memoryData['dingdan']['max']
                    ws['C5'] = self.memoryData['dingdan']['min']
                    ws['D5'] = self.memoryData['dingdan']['anv']
                    ws['E5'] = self.memoryData['dingdan']['top']
                    ws['F5'] = self.memoryData['dingdan']['60M']
                    ws['G5'] = self.memoryData['dingdan']['80M']
                    ws['H5'] = self.memoryData['dingdan']['100M']
                    ws['I5'] = self.memoryData['dingdan']['rst']
                    ws['B6'] = self.memoryData['denglu']['max']
                    ws['C6'] = self.memoryData['denglu']['min']
                    ws['D6'] = self.memoryData['denglu']['anv']
                    ws['E6'] = self.memoryData['denglu']['top']
                    ws['F6'] = self.memoryData['denglu']['60M']
                    ws['G6'] = self.memoryData['denglu']['80M']
                    ws['H6'] = self.memoryData['denglu']['100M']
                    ws['I6'] = self.memoryData['denglu']['rst']
                elif sheet_name == u'上行速率':
                    ws['B3'] = self.txData['dianying']['max']
                    ws['C3'] = self.txData['dianying']['min']
                    ws['D3'] = self.txData['dianying']['anv']
                    ws['E3'] = self.txData['dianying']['top']
                    ws['F3'] = self.txData['dianying']['10K']
                    ws['G3'] = self.txData['dianying']['20K']
                    ws['H3'] = self.txData['dianying']['40K']
                    ws['I3'] = self.txData['dianying']['60K']
                    ws['J3'] = self.txData['dianying']['80K']
                    ws['K3'] = self.txData['dianying']['100K']
                    ws['L3'] = self.txData['dianying']['rst']
                    ws['B4'] = self.txData['meishi']['max']
                    ws['C4'] = self.txData['meishi']['min']
                    ws['D4'] = self.txData['meishi']['anv']
                    ws['E4'] = self.txData['meishi']['top']
                    ws['F4'] = self.txData['meishi']['10K']
                    ws['G4'] = self.txData['meishi']['20K']
                    ws['H4'] = self.txData['meishi']['40K']
                    ws['I4'] = self.txData['meishi']['60K']
                    ws['J4'] = self.txData['meishi']['80K']
                    ws['K4'] = self.txData['meishi']['100K']
                    ws['L4'] = self.txData['meishi']['rst']
                    ws['B5'] = self.txData['dingdan']['max']
                    ws['C5'] = self.txData['dingdan']['min']
                    ws['D5'] = self.txData['dingdan']['anv']
                    ws['E5'] = self.txData['dingdan']['top']
                    ws['F5'] = self.txData['dingdan']['10K']
                    ws['G5'] = self.txData['dingdan']['20K']
                    ws['H5'] = self.txData['dingdan']['40K']
                    ws['I5'] = self.txData['dingdan']['60K']
                    ws['J5'] = self.txData['dingdan']['80K']
                    ws['K5'] = self.txData['dingdan']['100K']
                    ws['L5'] = self.txData['dingdan']['rst']
                    ws['B6'] = self.txData['denglu']['max']
                    ws['C6'] = self.txData['denglu']['min']
                    ws['D6'] = self.txData['denglu']['anv']
                    ws['E6'] = self.txData['denglu']['top']
                    ws['F6'] = self.txData['denglu']['10K']
                    ws['G6'] = self.txData['denglu']['20K']
                    ws['H6'] = self.txData['denglu']['40K']
                    ws['I6'] = self.txData['denglu']['60K']
                    ws['J6'] = self.txData['denglu']['80K']
                    ws['K6'] = self.txData['denglu']['100K']
                    ws['L6'] = self.txData['denglu']['rst']
                elif sheet_name == u'下行速率':
                    ws['B3'] = self.rxData['dianying']['max']
                    ws['C3'] = self.rxData['dianying']['min']
                    ws['D3'] = self.rxData['dianying']['anv']
                    ws['E3'] = self.rxData['dianying']['top']
                    ws['F3'] = self.rxData['dianying']['20K']
                    ws['G3'] = self.rxData['dianying']['40K']
                    ws['H3'] = self.rxData['dianying']['60K']
                    ws['I3'] = self.rxData['dianying']['80K']
                    ws['J3'] = self.rxData['dianying']['100K']
                    ws['K3'] = self.rxData['dianying']['rst']
                    ws['B4'] = self.rxData['meishi']['max']
                    ws['C4'] = self.rxData['meishi']['min']
                    ws['D4'] = self.rxData['meishi']['anv']
                    ws['E4'] = self.rxData['meishi']['top']
                    ws['F4'] = self.rxData['meishi']['20K']
                    ws['G4'] = self.rxData['meishi']['40K']
                    ws['H4'] = self.rxData['meishi']['60K']
                    ws['I4'] = self.rxData['meishi']['80K']
                    ws['J4'] = self.rxData['meishi']['100K']
                    ws['K4'] = self.rxData['meishi']['rst']
                    ws['B5'] = self.rxData['dingdan']['max']
                    ws['C5'] = self.rxData['dingdan']['min']
                    ws['D5'] = self.rxData['dingdan']['anv']
                    ws['E5'] = self.rxData['dingdan']['top']
                    ws['F5'] = self.rxData['dingdan']['20K']
                    ws['G5'] = self.rxData['dingdan']['40K']
                    ws['H5'] = self.rxData['dingdan']['60K']
                    ws['I5'] = self.rxData['dingdan']['80K']
                    ws['J5'] = self.rxData['dingdan']['100K']
                    ws['K5'] = self.rxData['dingdan']['rst']
                    ws['B6'] = self.rxData['denglu']['max']
                    ws['C6'] = self.rxData['denglu']['min']
                    ws['D6'] = self.rxData['denglu']['anv']
                    ws['E6'] = self.rxData['denglu']['top']
                    ws['F6'] = self.rxData['denglu']['20K']
                    ws['G6'] = self.rxData['denglu']['40K']
                    ws['H6'] = self.rxData['denglu']['60K']
                    ws['I6'] = self.rxData['denglu']['80K']
                    ws['J6'] = self.rxData['denglu']['100K']
                    ws['K6'] = self.rxData['denglu']['rst']
                elif sheet_name == u'电池温度':
                    ws['B3'] = self.batteryTemperatureData['dianying']['max']
                    ws['C3'] = self.batteryTemperatureData['dianying']['min']
                    ws['D3'] = self.batteryTemperatureData['dianying']['anv']
                    ws['E3'] = self.batteryTemperatureData['dianying']['top']
                    ws['F3'] = self.batteryTemperatureData['dianying']['60℃']
                    ws['G3'] = self.batteryTemperatureData['dianying']['80℃']
                    ws['H3'] = self.batteryTemperatureData['dianying']['rst']
                    ws['B4'] = self.batteryTemperatureData['meishi']['max']
                    ws['C4'] = self.batteryTemperatureData['meishi']['min']
                    ws['D4'] = self.batteryTemperatureData['meishi']['anv']
                    ws['E4'] = self.batteryTemperatureData['meishi']['top']
                    ws['F4'] = self.batteryTemperatureData['meishi']['60℃']
                    ws['G4'] = self.batteryTemperatureData['meishi']['80℃']
                    ws['H4'] = self.batteryTemperatureData['meishi']['rst']
                    ws['B5'] = self.batteryTemperatureData['dingdan']['max']
                    ws['C5'] = self.batteryTemperatureData['dingdan']['min']
                    ws['D5'] = self.batteryTemperatureData['dingdan']['anv']
                    ws['E5'] = self.batteryTemperatureData['dingdan']['top']
                    ws['F5'] = self.batteryTemperatureData['dingdan']['60℃']
                    ws['G5'] = self.batteryTemperatureData['dingdan']['80℃']
                    ws['H5'] = self.batteryTemperatureData['dingdan']['rst']
                    ws['B6'] = self.batteryTemperatureData['denglu']['max']
                    ws['C6'] = self.batteryTemperatureData['denglu']['min']
                    ws['D6'] = self.batteryTemperatureData['denglu']['anv']
                    ws['E6'] = self.batteryTemperatureData['denglu']['top']
                    ws['F6'] = self.batteryTemperatureData['denglu']['60℃']
                    ws['G6'] = self.batteryTemperatureData['denglu']['80℃']
                    ws['H6'] = self.batteryTemperatureData['denglu']['rst']
                elif sheet_name == u'FPS 性能':
                    ws['B2'] = self.fpsData['max']
                    ws['B3'] = self.fpsData['min']
                    ws['B4'] = self.fpsData['anv']
                    ws['B5'] = self.fpsData['16ms']
                    ws['B6'] = self.fpsData['20ms']
                    ws['B7'] = self.fpsData['rst']
                elif sheet_name == u'冷启动性能':
                    ws['B2'] = self.coldBootData['num']
                    ws['B3'] = self.coldBootData['max']
                    ws['B4'] = self.coldBootData['min']
                    ws['B5'] = self.coldBootData['anv']
                    ws['B6'] = self.coldBootData['top']
                    ws['B7'] = self.coldBootData['rst']
                elif sheet_name == u'热启动性能':
                    ws['B2'] = self.warmBootData['num']
                    ws['B3'] = self.warmBootData['max']
                    ws['B4'] = self.warmBootData['min']
                    ws['B5'] = self.warmBootData['anv']
                    ws['B6'] = self.warmBootData['top']
                    ws['B7'] = self.warmBootData['rst']
            xlsFile = os.path.join(self.attachmentPath, u'飞凡竞品性能评测报告(%s).xlsx' % time.strftime("%Y%m%d"))
            wb.save(xlsFile)
            if os.path.exists(file):
                os.remove(file)
        except:
            print("no sheet in %s named %s" % file,sheet_name)

class SendMail(object):
    def __init__(self):
        pass

    def mail(self):
        pass

def parse_command():
    '''
    解析日志路径命令行参数
    '''
    parser = argparse.ArgumentParser()
    parser.add_argument('-l', '--log_path', action='store', default='.',
                        dest='log_path', help='Setup log path, default is current execution directory.')
    args = parser.parse_args()
    rsPath = os.path.abspath(args.log_path)
    return rsPath

if __name__ == "__main__":
    rspath = parse_command()
    handler = Handler()
    handler.handle(rspath)
