import os
import time
import datetime
import shutil
import zipfile 
import openpyxl
from tools.utility.constants import *


ERROR = 0
CRASH = 0


class DataHandler(object):
    def __init__(self):
        self.rsPath = ''
        self.testCase = ''

    def handle(self, rsPath):
        self.rsPath = rsPath
        self.reportPath = os.path.join(self.rsPath, 'attachment')
        self.logPath = os.path.join(self.reportPath, 'log')

        self._summaryLogData()
        self._addZipFile()

    def getLogData(self,testCase):
        self.testCase = testCase
        return self._parserLogData()

    #把Log文件信息打包 
    def _addZipFile(self):
        zipf = zipfile.ZipFile('log.zip', 'w')
        pre_len = len(os.path.dirname(self.logPath))
        for parent, dirnames, filenames in os.walk(self.logPath):
            for filename in filenames:
                pathfile = os.path.join(parent, filename)
                arcname = pathfile[pre_len:].strip(os.path.sep)
                zipf.write(pathfile, arcname)
        zipf.close()
        zipFileName = os.path.join(self.reportPath, 'log.zip')
        shutil.move("log.zip", zipFileName)

    def _summaryLogData(self):
        if not os.path.exists(self.logPath):
            os.makedirs(self.logPath)

        print("debug logpath %s " % self.logPath)
        cmdCP = "cp -R %s/[0-9] %s" % (self.rsPath, self.logPath)
        print("debug cmd cp %s " % cmdCP)
        os.system(cmdCP)

        tmpFile = os.path.join(self.logPath, 'tmp.txt')
        cmdFind = 'find %s -type d -name "screenshot" > %s' % (self.logPath, tmpFile)
        #print("debug cmd cmdFind %s " % cmdFind)
        os.system(cmdFind)
        rmPath = open(tmpFile, 'r')
        while 1:
            line = rmPath.readline()
            if not line:
                break
            else:
                cmdRM = "rm -rf %s" % (line)
                os.system(cmdRM)
        if os.path.exists(tmpFile):
            os.remove(tmpFile)

    def _parserLogData(self):
        i = 0
        caseErrorInfo = {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]}
        tmpFile = os.path.join(self.logPath, '%s.txt') % self.testCase
        cmdFind = 'find %s -name "%s*.log" > %s' % (self.logPath, self.testCase, tmpFile)
        os.system(cmdFind)
        print("debug cmdFind  %s " % cmdFind)
        if os.path.exists(tmpFile):
            logCasePath = open(tmpFile, 'r')
            logPaths = logCasePath.readlines()
            logCasePath.close()
            if logPaths != []:
                for logPath in logPaths:
                    caseLogFile = logPath[:-1]
                    errorFilePath = caseLogFile.find('log')
                    if os.path.exists(caseLogFile):
                        #print("debug caseLogfile  %s " % caseLogFile)
                        logInfo = open(caseLogFile, 'r')
                        logLines = logInfo.readlines()
                        logInfo.close()
                        for logLine in logLines:
                            i += 1
                            if logLine.find("Error") != -1:
                                #print(logLine)
                                #过滤重复的error
                                #print("debug len(caseErrorInfo['ERRORLOG'] %s " % (len(caseErrorInfo['ERRORLOG'])))

                                # if(len(caseErrorInfo['ERRORLOG']) > 2):
                                #     #print("debug line %s " % logLine)
                                #     #print("debug logLine %s " % logLine)
                                #     str_filter = logLine.split(':')[3]
                                    #print("debug str_filter %s " % str_filter)
                                    #pass
                                #     for temp_str in caseErrorInfo['ERRORLOG']:
                                #         #如果已经保存的errlog中包含str_filter,则不再向其中添加
                                #         print("debug filter!!!!!!!!")
                                #         if temp_str.find(str_filter) == -1:
                                #             caseErrorInfo['ERROR'] += 1
                                #             caseErrorInfo['ERRORLOG'].append(logLine)
                                #             caseErrorInfo['ERRORNUM'].append(i)
                                #             caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])
                                # else:
                                caseErrorInfo['ERROR'] += 1
                                caseErrorInfo['ERRORLOG'].append(logLine)
                                caseErrorInfo['ERRORNUM'].append(i)
                                caseErrorInfo['ERRORFILE'].append(caseLogFile[errorFilePath:])

        print("error number %s " % caseErrorInfo['ERROR'])
        if os.path.exists(tmpFile):
            os.remove(tmpFile)
        return caseErrorInfo


class Handler(object):
    def __init__(self, deviceType):
        self.rsPath = ''
        self.reportPath = ''
        self.workbook = ''
        self.dataLength = 0
        self.dataList = {u'quanchengsousuo': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'gouwuzhongxin': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'meishihui': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangsousuo': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangzhaodian': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangpaidui': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangtingche': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'guangchangmaidan': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'wodedenglu': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]},
                         u'wodetuichu': {'ERROR': 0, 'CRASH': 0, 'ERRORLOG':[], 'ERRORNUM':[], 'ERRORFILE':[]}}
        if deviceType == 'Android':
            from configs.androidConfig import appVersion, phoneVersion, buildVersion, deviceID, deviceNet
        elif deviceType == 'IOS':
            from configs.iosConfig import appVersion, phoneVersion, buildVersion, deviceID, deviceNet
        else:
            raise

        self.deviceType = deviceType
        self.appVersion = appVersion
        self.phoneVersion = phoneVersion
        self.buildVersion = buildVersion
        self.deviceID = deviceID
        self.deviceNet = deviceNet

    def handle(self, rsPath):
        self.rsPath = rsPath
        self._mkReportDir()
        try:
            resourcesDirectory = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/stability_test/"
            xlsFileTemplate = os.path.join(resourcesDirectory, u'templateStability.xlsx')
            shutil.copy(xlsFileTemplate,self.reportPath)
            dataHandler = DataHandler()
            dataHandler.handle(rsPath=self.rsPath)
            for testCase in (STABILITY_CASE_FOLDER_LIST.values()):
                print("debug test case %s " % testCase)
                caseErrorInfo = dataHandler.getLogData(testCase=testCase)
                self.dataList[testCase]['ERROR'] = caseErrorInfo['ERROR']
                self.dataList[testCase]['CRASH'] = caseErrorInfo['CRASH']
                self.dataList[testCase]['ERRORLOG'] = caseErrorInfo['ERRORLOG']
                self.dataList[testCase]['ERRORNUM'] = caseErrorInfo['ERRORNUM']
                self.dataList[testCase]['ERRORFILE'] = caseErrorInfo['ERRORFILE']


            global ERROR
            ERROR = self.dataList['quanchengsousuo']['ERROR'] + self.dataList['gouwuzhongxin']['ERROR'] +\
                        self.dataList['meishihui']['ERROR'] + self.dataList['guangchangsousuo']['ERROR'] + \
                        self.dataList['guangchangzhaodian']['ERROR'] + self.dataList['guangchangpaidui']['ERROR'] +\
                        self.dataList['guangchangtingche']['ERROR'] + self.dataList['guangchangmaidan']['ERROR'] +\
                        self.dataList['wodedenglu']['ERROR'] + self.dataList['wodetuichu']['ERROR']
            global CRASH
            CRASH = self.dataList['quanchengsousuo']['CRASH'] + self.dataList['gouwuzhongxin']['CRASH'] +\
                        self.dataList['meishihui']['CRASH'] + self.dataList['guangchangsousuo']['CRASH'] +\
                        self.dataList['guangchangzhaodian']['CRASH'] + self.dataList['guangchangpaidui']['CRASH'] +\
                        self.dataList['guangchangtingche']['CRASH'] + self.dataList['guangchangmaidan']['CRASH'] +\
                        self.dataList['wodedenglu']['CRASH'] + self.dataList['wodetuichu']['CRASH']

            xlsFile = os.path.join(self.reportPath, u'templateStability.xlsx')
        except Exception as e:
            print(e)
        finally:
            self._writeExcel(xlsFile)

    def _mkReportDir(self):
        '''
        创建报告目录, 包含excel报告,log压缩包
        '''
        self.reportPath = os.path.join(self.rsPath, 'attachment')
        if os.path.exists(self.reportPath):
            timestamp = os.path.getmtime(self.reportPath)
            date = datetime.datetime.fromtimestamp(timestamp).strftime('%Y_%m_%d_%H_%M_%S')
            newDirName = 'report_' + date

            newLogDir = os.path.join(self.rsPath, newDirName)
            try:
                os.renames(self.reportPath, newLogDir)
            except Exception as e:
                raise IOError('Modify the [%s] directory name failed with following error: \n'
                              '%s' % (self.reportPath, e))

        try:
            os.makedirs(self.reportPath)
        except Exception as e:
            raise IOError('Create the [%s] directory failed with following error: \n'
                          '%s' % (self.reportPath, e))

    def _writeExcel(self, file='file.xls'):
        try:
            print("debug file %s" % file)
            wb = openpyxl.load_workbook(file)
            for sheet_name in STABILITY_REPORT_SHEET:
                print("sheet name %s " % sheet_name)
                ws = wb.get_sheet_by_name(sheet_name)
                if sheet_name == u'测试环境':
                    ws['C4'] = self.phoneVersion
                    ws['C5'] = self.buildVersion
                    ws['C9'] = self.appVersion
                    ws['A13'] = self.deviceNet
                elif sheet_name == u'全城搜索':
                    print("debug write data to 全城搜索 sheet")
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['quanchengsousuo']['ERROR']
                    ws['C6'] = self.dataList['quanchengsousuo']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']

                    if len(self.dataList['quanchengsousuo']['ERRORLOG']) != 0:
                        print("debug write log to 全城搜索 sheet")
                        for i in range(len(self.dataList['quanchengsousuo']['ERRORLOG'])):
                            log_quanchengsousuo = 'A' + str(i + 10)
                            line_quanchengsousuo = 'E' + str(i + 10)
                            message_quanchengsousuo = 'F' + str(i + 10)
                            ws[line_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORNUM'][i]
                            ws[message_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORLOG'][i]
                            ws[log_quanchengsousuo] = self.dataList['quanchengsousuo']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'购物中心':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['gouwuzhongxin']['ERROR']
                    ws['C6'] = self.dataList['gouwuzhongxin']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']

                    if len(self.dataList['gouwuzhongxin']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['gouwuzhongxin']['ERRORLOG'])):
                            log_gouwuzhongxin = 'A' + str(i + 10)
                            line_gouwuzhongxin = 'E' + str(i + 10)
                            message_gouwuzhongxin = 'F' + str(i + 10)
                            ws[line_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORNUM'][i]
                            ws[message_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORLOG'][i]
                            ws[log_gouwuzhongxin] = self.dataList['gouwuzhongxin']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'美食汇':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['meishihui']['ERROR']
                    ws['C6'] = self.dataList['meishihui']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['meishihui']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['meishihui']['ERRORLOG'])):
                            log_meishihui = 'A' + str(i + 10)
                            line_meishihui = 'E' + str(i + 10)
                            message_meishihui = 'F' + str(i + 10)
                            ws[line_meishihui] = self.dataList['meishihui']['ERRORNUM'][i]
                            ws[message_meishihui] = self.dataList['meishihui']['ERRORLOG'][i]
                            ws[log_meishihui] = self.dataList['meishihui']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场搜索':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangsousuo']['ERROR']
                    ws['C6'] = self.dataList['guangchangsousuo']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['guangchangsousuo']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangsousuo']['ERRORLOG'])):
                            log_guangchangsousuo = 'A' + str(i + 10)
                            line_guangchangsousuo = 'E' + str(i + 10)
                            message_guangchangsousuo = 'F' + str(i + 10)
                            ws[line_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORNUM'][i]
                            ws[message_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORLOG'][i]
                            ws[log_guangchangsousuo] = self.dataList['guangchangsousuo']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场找店':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangzhaodian']['ERROR']
                    ws['C6'] = self.dataList['guangchangzhaodian']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['guangchangzhaodian']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangzhaodian']['ERRORLOG'])):
                            log_guangchangzhaodian = 'A' + str(i + 10)
                            line_guangchangzhaodian = 'E' + str(i + 10)
                            message_guangchangzhaodian = 'F' + str(i + 10)
                            ws[line_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORNUM'][i]
                            ws[message_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORLOG'][i]
                            ws[log_guangchangzhaodian] = self.dataList['guangchangzhaodian']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场排队':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangpaidui']['ERROR']
                    ws['C6'] = self.dataList['guangchangpaidui']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['guangchangpaidui']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangpaidui']['ERRORLOG'])):
                            log_guangchangpaidui = 'A' + str(i + 10)
                            line_guangchangpaidui = 'E' + str(i + 10)
                            message_guangchangpaidui = 'F' + str(i + 10)
                            ws[line_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORNUM'][i]
                            ws[message_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORLOG'][i]
                            ws[log_guangchangpaidui] = self.dataList['guangchangpaidui']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场停车':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangtingche']['ERROR']
                    ws['C6'] = self.dataList['guangchangtingche']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['guangchangtingche']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangtingche']['ERRORLOG'])):
                            log_guangchangtingche = 'A' + str(i + 10)
                            line_guangchangtingche = 'E' + str(i + 10)
                            message_guangchangtingche = 'F' + str(i + 10)
                            ws[line_guangchangtingche] = self.dataList['guangchangtingche']['ERRORNUM'][i]
                            ws[message_guangchangtingche] = self.dataList['guangchangtingche']['ERRORLOG'][i]
                            ws[log_guangchangtingche] = self.dataList['guangchangtingche']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'广场买单':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['guangchangmaidan']['ERROR']
                    ws['C6'] = self.dataList['guangchangmaidan']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['guangchangmaidan']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['guangchangmaidan']['ERRORLOG'])):
                            log_guangchangmaidan = 'A' + str(i + 10)
                            line_guangchangmaidan = 'E' + str(i + 10)
                            message_guangchangmaidan = 'F' + str(i + 10)
                            ws[line_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORNUM'][i]
                            ws[message_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORLOG'][i]
                            ws[log_guangchangmaidan] = self.dataList['guangchangmaidan']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'我的登录':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['wodedenglu']['ERROR']
                    ws['C6'] = self.dataList['wodedenglu']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['wodedenglu']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['wodedenglu']['ERRORLOG'])):
                            log_wodedenglu = 'A' + str(i + 10)
                            line_wodedenglu = 'E' + str(i + 10)
                            message_wodedenglu = 'F' + str(i + 10)
                            ws[line_wodedenglu] = self.dataList['wodedenglu']['ERRORNUM'][i]
                            ws[message_wodedenglu] = self.dataList['wodedenglu']['ERRORLOG'][i]
                            ws[log_wodedenglu] = self.dataList['wodedenglu']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
                elif sheet_name == u'我的退出':
                    ws['C4'] = INSIDELOOPNUM * OUTLOOPNUM
                    ws['C5'] = self.dataList['wodetuichu']['ERROR']
                    ws['C6'] = self.dataList['wodetuichu']['ERROR']
                    ws['C7'] = self.dataList['quanchengsousuo']['CRASH']
                    if len(self.dataList['wodetuichu']['ERRORLOG']) != 0:
                        for i in range(len(self.dataList['wodetuichu']['ERRORLOG'])):
                            log_wodetuichu = 'A' + str(i + 10)
                            line_wodetuichu = 'E' + str(i + 10)
                            message_wodetuichu = 'F' + str(i + 10)
                            ws[line_wodetuichu] = self.dataList['wodetuichu']['ERRORNUM'][i]
                            ws[message_wodetuichu] = self.dataList['wodetuichu']['ERRORLOG'][i]
                            ws[log_wodetuichu] = self.dataList['wodetuichu']['ERRORFILE'][i]
                    else:
                        ws['A11'] = '-'
                        ws['C11'] = '-'
                        ws['D11'] = '-'
            xlsFile = os.path.join(self.reportPath, u'飞凡APP重点功能压力测试报告(%s).xlsx' % time.strftime("%Y%m%d"))
            wb.save(xlsFile)
            if os.path.exists(file):
                os.remove(file)
        except:
            print("no sheet in %s named %s" % (file, sheet_name))


if __name__ == "__main__":
    handler = Handler('IOS')
    resultsPath = "%s/report/stability/%s/%s" % ("/Users/auto/workspace_pycharm/autotest", '20170123', '1')
    if os.path.exists(resultsPath):
        handler.handle(resultsPath)
